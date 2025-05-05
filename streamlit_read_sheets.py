import streamlit as st
import openpyxl
import pandas as pd
import util.script_generator as sg
import util.df_utlis as dfutils
import util.connect_to_cs as cs

st.title("View the workbook")

st.markdown(
    """<style>
        .element-container:nth-of-type(3) button {
            height: 3em;
        }
        </style>""",
    unsafe_allow_html=True,
)

# Initialize session state
if 'sql_cmd' not in st.session_state:
    st.session_state.sql_cmd = ''
if 'sheet' not in st.session_state:
    st.session_state.sheet=''



def execute_sql(sql_script):
    print(sql_script)
    st.write(sql_script)
    response = cs.create_core_tables(sql_script)
    return response


def group_by_table(df):
    d= dict(df.groupby('Table Name').apply(list))
    st.write(d)
    return d

def create_script(df,sheet):
    returned_script=''
    match sheet:
        case "CORE tables" :
            st.write(f"{sheet} in process.")  # Match weekends
            returned_script = sg.core_tables_script(df)
        case "System":
            st.write(f"{sheet} in process.")  # Match sheet
            returned_script = sg.process_system_tab(df)
        case "Stream":
            st.write(f"{sheet} in process.")  
            returned_script = sg.process_stream_tab(df)
        case "STG Tables":
            st.write(f"{sheet} in process.")  # Match sheet
            st.write(df.columns)
        case _:
            st.write("Thatis default.")  # Default case

    #print(df)
    st.write(returned_script)   
    return returned_script


def check_difference(df1, df2):
    difference = df1[df1!=df2]
    st.write(difference) 

def dataframe_with_selections(df):
    df_with_selections = df.copy()
    df_with_selections.insert(0, "Select", False)

    # Get dataframe row-selections from user with st.data_editor
    edited_df = st.data_editor(
        df_with_selections,
        hide_index=True,
        column_config={"Select": st.column_config.CheckboxColumn(required=True)},
        disabled=df.columns,
    )

    # Filter the dataframe using the temporary column, then drop the column
    selected_rows = edited_df[edited_df.Select]
    return selected_rows.drop('Select', axis=1)

data_file = st.sidebar.file_uploader("Upload Excel file",type=['xlsx'])  

if data_file:
    file_details = {
        "Filename":data_file.name,
        "FileType":data_file.type,
        "FileSize":data_file.size}

    wb = openpyxl.load_workbook(data_file)

    ## Show Excel file
    st.sidebar.subheader("File details:")
    st.sidebar.json(file_details,expanded=False)
    st.sidebar.markdown("----")

    ## Select sheet
    sheet_selector = st.sidebar.selectbox("Select sheet:",wb.sheetnames)     
    df = pd.read_excel(data_file,sheet_selector)
    st.markdown(f"### Currently Selected: `{sheet_selector}`")
    st.write(df)
    if sheet_selector =='CORE tables':
        tables_df = dfutils.df_groupBy(df, 'Table Name')
        st.write(tables_df)



    mycomment = '''
    ## Do something after a button
    doLogic_btn = st.button("➕")
    if doLogic_btn:
        df2 = df.sum().transpose()
        st.write(df2)

        # Do something more after the previous button
        # >> But this will fail because the button will go back to _False_ 
        # >> so nothing will be shown afterwards
        another_btn = st.checkbox("Another +")
        if another_btn:
            df3 = df2.sum()
            st.write(df3)

            '''
    
data_file2 = st.sidebar.file_uploader("Upload the previous Excel file ",type=['xlsx'])  

if data_file2:
    file_details2 = {
        "Filename":data_file2.name,
        "FileType":data_file2.type,
        "FileSize":data_file2.size}

    wb2 = openpyxl.load_workbook(data_file2)
    print(wb2)
    ## Show Excel file
    st.sidebar.subheader("File 2 details:")
    st.sidebar.json(file_details2,expanded=False)
    st.sidebar.markdown("----")

    ## Select sheet
    sheet_selector2 = st.sidebar.selectbox("Select sheet from the second file:",wb2.sheetnames)     
    df2 = pd.read_excel(data_file2,sheet_selector2)

    st.markdown(f"### Currently Selected: `{sheet_selector2}`")
    st.write(df2)
    st.session_state.sheet=sheet_selector2
    print(st.session_state.sheet)
    my_comment = '''
    event = st.dataframe(
        df2,
        use_container_width=True,
        hide_index=True,
        on_select="rerun",
        selection_mode="multi-row",
        )
        '''
    st.write("Select an item to convert to SQL:")

    if sheet_selector2 =='CORE tables':
        tables_df2 = dfutils.df_groupBy(df2, 'Table Name')
        #st.write(tables_df2)
        selection = dataframe_with_selections(tables_df2)
        displayed_selection =selection.copy()
        displayed_selection['attr'] = displayed_selection['attr'].apply(lambda x: str(x))
        
    else :
        selection = dataframe_with_selections(df2)
        displayed_selection =selection

    st.write("Your selection:")
    st.write(displayed_selection)

    

    if st.button('compare'):
        check_difference(df, df2)

    if st.button('generate sql'):
        rs = create_script(selection, st.session_state.sheet)
        st.session_state.sql_cmd = rs
        #selection['sql_script']=rs
        #st.dataframe(selection,column_config={'Table Name': 'Table',
                  #'sql_script': 'SQL Script'}  , hide_index=True )
    if st.button('execute sql'):
        rs= st.session_state.sql_cmd 
        print('-------------------------------------------------', type(rs))
        st.write(st.session_state.sql_cmd )
        r=execute_sql(rs)
        st.write(r)
        
     
   


